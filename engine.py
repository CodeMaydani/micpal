"""
engine.py
Core extraction / generation / insertion logic for the מיכפל salary
import automation.

Pure functions only -- no printing, no hardcoded paths. Both
gen_company.py (CLI) and app.py (Streamlit UI) import from here so
behavior is identical between the two front ends.
"""

import datetime
import glob
import os
import re
import struct

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def col_letter(n):
    """Convert 1-based column index to Excel letter(s): 1->A, 27->AA, etc."""
    r = ""
    while n:
        n, x = divmod(n - 1, 26)
        r = chr(65 + x) + r
    return r


def valid_israeli_id(n):
    """Luhn-variant check for Israeli 9-digit ID numbers."""
    s = str(n).zfill(9)
    total = 0
    for i, c in enumerate(s):
        d = int(c) * (1 if i % 2 == 0 else 2)
        total += d - 9 if d > 9 else d
    return total % 10 == 0


# Excel/XML 1.0 forbids most control characters in cell strings. Allowed:
# tab (\x09), newline (\x0A), carriage return (\x0D), and the printable
# ranges above \x20 (excluding the surrogate/illegal blocks). openpyxl
# raises IllegalCharacterError on anything outside this set.
_ILLEGAL_XML_CHARS_RE = re.compile(
    "[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x84\x86-\x9f\ufdd0-\ufddf\ufffe\uffff]"
)


def sanitize_text(s):
    """Strip characters that are illegal in Excel/XML cell values."""
    return _ILLEGAL_XML_CHARS_RE.sub("", s).strip()


# ---------------------------------------------------------------------------
# Company discovery / template inspection
# ---------------------------------------------------------------------------


def discover_companies(data_dir):
    """
    Find company codes that have BOTH a Q8MIFL26.[company] and a
    Q8OVDM26.[company] file in data_dir.

    Returns a sorted list of company-code strings exactly as they appear
    in the filename suffix (leading zeros preserved).
    """
    mifl = {
        os.path.basename(p).split(".", 1)[1]
        for p in glob.glob(os.path.join(data_dir, "Q8MIFL26.*"))
    }
    ovdm = {
        os.path.basename(p).split(".", 1)[1]
        for p in glob.glob(os.path.join(data_dir, "Q8OVDM26.*"))
    }
    return sorted(mifl & ovdm)


def list_existing_templates(srgl_path):
    """
    Parse @N lines out of Q8SRGL26.000 and return the template names.
    Used for duplicate-name warnings before generating a new template.
    """
    with open(srgl_path, "rb") as f:
        content = f.read()
    text = content.decode("iso-8859-8", errors="replace")
    return [name.strip() for name in re.findall(r"^@N(.+)$", text, re.MULTILINE)]


# ---------------------------------------------------------------------------
# Binary extraction
# ---------------------------------------------------------------------------


def extract_components(path):
    """
    Read salary components from Q8MIFL26.[company].
    Returns (results, skipped_count).
      results       -- list of (rechiv_extracted, name, kod_mahk)
      skipped_count -- candidate records that looked like real components
                       but failed a filter (diagnostic only)

    NOTE: actual מיכפל code = rechiv_extracted - 1.

    NUMBERING FIX (was: rechiv_extracted = tail[10], the "-1 offset
    invariant"): tail[10] of a record is actually the NEXT record's
    number, not this one's -- it only looked like "current+1" because
    that holds whenever component numbering has no gaps. Verified against
    the live מיכפל UI across companies 003/004/083: the correct number is
    the PREVIOUS record's tail[10], read directly as
    data[name_offset - 19]. The first record has no predecessor, so it
    falls back to tail[10] itself (always correct there, since משכורת is
    always record 0).

    NOISE-RECORD GUARD (new): company 003's file contains a few garbled
    pseudo-records past the real component list that satisfied the old
    filters by coincidence. Excluded via looks_like_real_component_name()
    -- a real name is composed almost entirely of Hebrew letters, digits,
    and common punctuation; verified with zero false rejections across
    189 known-real components and 3/3 known garbage records caught.
    """
    data = open(path, "rb").read()
    TAIL_SIZE = 29
    START = 0x4E2E
    MAX_REXIVIM = 252  # confirmed in מיכפל's own UI: max רכיבי שכר per company
    results = []
    skipped = 0
    offset = START
    first = True
    while offset < len(data) - TAIL_SIZE:
        try:
            name_end = data.index(b"\x00", offset)
        except ValueError:
            break
        raw_name = data[offset:name_end].decode("iso-8859-8", errors="replace").strip()
        name = sanitize_text(raw_name)
        tail = data[name_end + 1 : name_end + 1 + TAIL_SIZE]
        if len(tail) < TAIL_SIZE:
            break
        t10 = tail[10]
        kod_mahk = tail[-1]

        # --- numbering fix: real number = previous record's tail[10] ---
        if first:
            rechiv_extracted = t10
        else:
            rechiv_extracted = data[offset - 19] + 1
        # -----------------------------------------------------------------

        actual = rechiv_extracted - 1
        if (
            name
            and t10 > 0
            and t10 < 200
            and 0 < actual <= MAX_REXIVIM
            and "\ufffd" not in name
            and name != "?" * len(name)
            and kod_mahk < 100
            and looks_like_real_component_name(name)  # noise-record guard
        ):
            results.append((rechiv_extracted, name, kod_mahk))
        elif raw_name and t10 > 0 and t10 < 200:
            skipped += 1
        first = False
        offset = name_end + 1 + TAIL_SIZE
    return results, skipped


_ALLOWED_PUNCT = set(" .,'\"%-~$()/+")
_MAX_NOISE_FRACTION = 0.15


# ---------------------------------------------------------------------------
# Deduction components (רכיבי ניכוי רשות) -- a SEPARATE family from salary
# ---------------------------------------------------------------------------
# These live in their own definition table in Q8MIFL26, distinct from the
# salary components (which start at 0x4e2e). The deduction table is a packed
# array of fixed 20-byte records starting at DEDUCTION_DEF_START:
#     bytes [0:10]   name, null-padded ISO-8859-8
#     bytes [10:14]  קוד מה"כ (4-char code, e.g. עאממ / מאהמ / מאשמ)
#     bytes [14:20]  padding (zero)
# The deduction NUMBER is simply the zero-based index in this table (verified
# against the live מיכפל dropdown: 0=מפרעה, 1=החזר הלואה, 2=חנות המפעל, ...).
# Terminated by the first record whose name is the '..........' placeholder
# (i.e. not a real Hebrew name).
DEDUCTION_DEF_START = 0x3F9D
DEDUCTION_DEF_SIZE = 20
MAX_DEDUCTIONS = 100  # generous guard; real tables are small (single digits)


def extract_deductions(path):
    """
    Parse the deduction-definition table from Q8MIFL26.[company].

    Returns a list of (deduction_number, name, kod_mahk) tuples, where
    deduction_number is the zero-based index (matches the מיכפל dropdown and
    the value stored per-employee; see read_prior_month's deduction region).
    kod_mahk is the 4-char code string. The list is in table order and stops
    at the first non-real-name record (the '..........' placeholder).
    """
    data = open(path, "rb").read()
    results = []
    for i in range(MAX_DEDUCTIONS):
        off = DEDUCTION_DEF_START + i * DEDUCTION_DEF_SIZE
        if off + DEDUCTION_DEF_SIZE > len(data):
            break
        raw_name = (
            data[off : off + 10]
            .rstrip(b"\x00")
            .decode("iso-8859-8", errors="replace")
            .strip()
        )
        name = sanitize_text(raw_name)
        if not name or "\ufffd" in name or not looks_like_real_component_name(name):
            break  # placeholder / end of table
        kod_mahk = (
            data[off + 10 : off + 14]
            .decode("iso-8859-8", errors="replace")
            .strip("\x00")
        )
        results.append((i, name, kod_mahk))
    return results


def looks_like_real_component_name(name):
    """
    Content-based filter distinguishing real component names from garbage
    pseudo-records found past the end of a company's real component
    section (e.g. company 003 near the documented 0xCCCA GetLength-clamp
    boundary). An offset-gap approach was tried first and rejected: the
    transition isn't one clean jump -- the parser walks through many
    small steps (50-220 bytes) of padding-as-pseudo-records before
    reaching the actual garbage, so no gap-size threshold reliably marks
    the boundary. This content-based check works instead: a real name is
    built almost entirely from Hebrew letters, digits, and a small set of
    common punctuation; garbage records are dominated by other characters
    even when 1-2 Hebrew letters appear by coincidental byte alignment.
    """
    if not name or len(name) > 40:
        return False
    has_hebrew = any("\u05d0" <= ch <= "\u05ea" for ch in name)
    if not has_hebrew:
        return False
    noise = sum(
        1
        for ch in name
        if not ("\u05d0" <= ch <= "\u05ea" or ch.isdigit() or ch in _ALLOWED_PUNCT)
    )
    return (noise / len(name)) <= _MAX_NOISE_FRACTION


# ---------------------------------------------------------------------------
# Employee status (קוד הפסקה) -- single byte per Q8OVDM26 block
# ---------------------------------------------------------------------------
# Found by diffing a known-inactive employee (ישראל ישראלי, shown in the
# מיכפל UI with קוד הפסקה == 1) against active employees across every block
# of a test company. Offset 788 (0x314) was the ONLY byte that read 0 for
# every active employee, 1 for the inactive one, AND a distinct third value
# (2) for the file's pre-allocated empty slots -- the signature of a real
# status enum rather than coincidental data. The inactive record also
# carries a nonzero value at 784-786 (an associated תאריך הפסקה) where active
# records are zero, corroborating the field.
#
# Validation rule for any new company: the count of blocks with a real name
# AND byte[788] == 0 must match the UI's active-employee count.
EMPLOYEE_BLOCK = 167936
KOD_HAFSAKA_OFFSET = 788
STATUS_ACTIVE = 0
STATUS_INACTIVE = 1
STATUS_UNUSED_SLOT = 2


# ---------------------------------------------------------------------------
# Prior-month values (carry-forward) -- packed component-line table per block
# ---------------------------------------------------------------------------
# Inside each EMPLOYEE_BLOCK, מיכפל stores that employee's actual per-component
# salary lines as a packed array of 22-byte slots beginning at SLOT_TABLE_OFFSET,
# in ascending-component-number order, terminated by the first all-zero slot.
# Decoded and validated against company 004 with a deliberate component-number
# GAP test (components 3/7/18, 1/20, 1/9/10 across three employees): byte [7] of
# each slot holds the REAL מיכפל component number (proven by the gaps -- a
# positional index would have read 1,2,3; it read the true numbers), and the
# quantity/price are int32 little-endian scaled x100 (agorot). 8/8 component
# lines decoded exactly.
#
# IMPORTANT JOIN NOTE: slot[7] equals col_map's `actual` directly (= the real
# מיכפל component number; rechiv_extracted - 1, with the משכורת block = 1), NOT
# actual + 1. build_excel()'s comp_to_cols relies on this to map a carried
# component to its Excel columns; an earlier +1 here shifted every value one
# component to the left (caught against company 004's real export).
#
# byte [9] is the per-component ברוטו/נטו flag (0xE1 -> ב ברוטו, 0xF0 -> נ נטו),
# sitting in a fixed frame: byte [8] = 0xE3 (ד) const, byte [10] = 0x01 const,
# byte [11] = 0x30 const. The flag is per (employee, component) and CAN differ
# within one employee (company 004, זורבבל: comp 3=ב, 7=נ, 18=ב), so it is read
# and carried per component into each component's own ב/נ column -- not a single
# shared column. Decoded generically from the ISO-8859-8 Hebrew range so an
# unforeseen flag (e.g. ג for גילום) passes through as its letter.
SLOT_TABLE_OFFSET = 6303  # first component-line slot within an EMPLOYEE_BLOCK
SLOT_SIZE = 22
SLOT_COMP_NUM_OFF = 7  # within slot: component number (real מיכפל #), 1 byte
SLOT_BN_OFF = 9  # within slot: ברוטו/נטו flag (0xE1=ב, 0xF0=נ), 1 byte
SLOT_QTY_OFF = 14  # within slot: quantity x100, int32 LE
SLOT_PRICE_OFF = 18  # within slot: price    x100, int32 LE
_SLOT_WALK_GUARD = 260  # never walk past a plausible slot count (252 + margin)


def _decode_bn(byte):
    """Decode a salary slot's ברוטו/נטו flag byte (slot byte [9]) to a single
    Hebrew character: ב (ברוטו) for 0xE1, נ (נטו) for 0xF0. Any non-Hebrew byte
    (including 0x00) returns "" so the cell is left blank. Decoded from the
    ISO-8859-8 Hebrew letter range rather than a fixed {ב,נ} map, so a flag we
    have not seen yet (e.g. ג for גילום) surfaces as its letter instead of being
    silently dropped."""
    if 0xE0 <= byte <= 0xFA:  # ISO-8859-8 Hebrew letters א..ת
        return bytes([byte]).decode("iso-8859-8")
    return ""


# Deduction values per employee (רכיבי ניכוי רשות). Stored in a SEPARATE,
# tighter table than salary: packed 5-byte records in a "mirror" region at a
# fixed offset, each [amount int32 LE x100][trailing byte], preceded by a 0x01
# flag. Deductions are PRICE-ONLY (no quantity) plus a month field (in the
# primary copy near offset 6252). The trailing byte holds the NEXT record's
# (deduction_number + 1) -- the same "next record's number" quirk as the salary
# table -- so the correct deduction number is the PREVIOUS record's trailing
# byte minus 1, with the first record = deduction 0. Validated against company
# 004 across two gap tests (indices 0,2,5 -> trailing 3,6,0; and 0,1,2 ->
# trailing 2,3,0). Reading the mirror region (clean 5-byte records) here; the
# month lives in the primary copy and is not needed for carry-forward.
DEDUCTION_MIRROR_OFFSET = 70631  # first deduction record within EMPLOYEE_BLOCK
DEDUCTION_REC_SIZE = 5
DEDUCTION_AMOUNT_OFF = 0  # within record: amount x100, int32 LE
DEDUCTION_TRAILING_OFF = 4  # within record: NEXT record's (number + 1)
_DEDUCTION_WALK_GUARD = 110


def read_prior_month_deductions(path):
    """
    Read prior-month per-employee deduction amounts from Q8OVDM26.[company].

    Returns { tz_str : { deduction_number (int) : amount (float) } }.
    deduction_number is the zero-based index matching extract_deductions()
    and the מיכפל dropdown. amount is whole-shekel float (x100 un-scaled).
    Employees with no deductions map to an empty dict (same convention as
    read_prior_month). Filtering mirrors read_prior_month exactly so keys
    line up with sheet rows.
    """
    data = open(path, "rb").read()
    total_blocks = len(data) // EMPLOYEE_BLOCK
    out = {}
    for b in range(total_blocks):
        block = data[b * EMPLOYEE_BLOCK : (b + 1) * EMPLOYEE_BLOCK]
        status = block[KOD_HAFSAKA_OFFSET]
        if status not in (STATUS_ACTIVE, STATUS_INACTIVE):
            continue
        end = block.index(b"\x00", 9)
        last_name = sanitize_text(
            block[9:end].decode("iso-8859-8", errors="replace").strip()
        )
        end2 = block.index(b"\x00", 29)
        first_name = sanitize_text(
            block[29:end2].decode("iso-8859-8", errors="replace").strip()
        )
        full_name = f"{first_name} {last_name}".strip()
        if not full_name or "\ufffd" in full_name:
            continue
        tz_raw = struct.unpack_from("<I", block, 59)[0]
        if tz_raw == 0 or not valid_israeli_id(tz_raw):
            continue
        out[str(tz_raw).zfill(9)] = _read_block_deductions(block)
    return out


def _read_block_deductions(block):
    """
    Parse one employee's deduction mirror region into
    {deduction_number: amount}. Uses the previous-record's trailing byte to
    recover each record's deduction number (the trailing byte holds the NEXT
    record's number+1; the first record is deduction 0).
    """
    out = {}
    prev_trailing = None
    for i in range(_DEDUCTION_WALK_GUARD):
        off = DEDUCTION_MIRROR_OFFSET + i * DEDUCTION_REC_SIZE
        if off + DEDUCTION_REC_SIZE > len(block):
            break
        amount_raw = struct.unpack_from("<i", block, off + DEDUCTION_AMOUNT_OFF)[0]
        trailing = block[off + DEDUCTION_TRAILING_OFF]
        if amount_raw == 0 and trailing == 0:
            break  # terminator
        if i == 0 or prev_trailing is None:
            ded_number = 0
        else:
            ded_number = prev_trailing - 1
        out[ded_number] = amount_raw / 100.0
        prev_trailing = trailing
    return out


def read_prior_month(path):
    """
    Read prior-month per-employee, per-component salary values out of a
    Q8OVDM26.[company] file, for carry-forward population of a new month's
    sheet.

    Returns a dict keyed by zero-padded תעודת זהות string:

        { tz_str : { component_number (int) : (qty, price, bn) } }

    qty and price are floats in whole-shekel units (the on-disk x100 agorot
    scaling is undone here); bn is the component's ברוטו/נטו flag as a single
    character ("ב" ברוטו / "נ" נטו, or "" if the slot carries no flag), read
    from slot byte [9]. component_number is the REAL מיכפל number from slot
    byte [7], which equals col_map's `actual` (= rechiv_extracted - 1, with
    the משכורת block = component 1) -- the key build_excel() joins on.

    Keyed by תז (not the raw int) so it joins directly against the employee
    rows build_excel() writes, which use the same zero-padded valid-תז
    string. Blocks with an invalid/zero תז are skipped (they can't be
    matched to a sheet row anyway). An employee present with no component
    lines maps to an empty dict {} -- distinct from an employee absent from
    the prior month (the tz simply isn't a key), so callers can tell
    "no activity last month" from "new hire".
    """
    data = open(path, "rb").read()
    total_blocks = len(data) // EMPLOYEE_BLOCK
    out = {}
    for b in range(total_blocks):
        block = data[b * EMPLOYEE_BLOCK : (b + 1) * EMPLOYEE_BLOCK]
        status = block[KOD_HAFSAKA_OFFSET]
        if status not in (STATUS_ACTIVE, STATUS_INACTIVE):
            continue  # unused slot / not a real employee

        # Mirror _read_employee_records()'s name filtering so this dict's
        # keys line up exactly with the employees that get sheet rows.
        # Otherwise a filtering mismatch could flag a real employee as a
        # "new hire" purely because the two readers disagreed on whether
        # the block counts.
        end = block.index(b"\x00", 9)
        last_name = sanitize_text(
            block[9:end].decode("iso-8859-8", errors="replace").strip()
        )
        end2 = block.index(b"\x00", 29)
        first_name = sanitize_text(
            block[29:end2].decode("iso-8859-8", errors="replace").strip()
        )
        full_name = f"{first_name} {last_name}".strip()
        if not full_name or "\ufffd" in full_name:
            continue  # nameless / garbage block

        tz_raw = struct.unpack_from("<I", block, 59)[0]
        # tz_raw == 0 passes the Luhn check (0 % 10 == 0) but is not a real
        # ID; reject it explicitly. A block whose תז is 0/invalid can't be
        # joined to a sheet row (those rows carry tz ""), so skip it -- this
        # keeps read_prior_month's keys to genuinely matchable employees.
        if tz_raw == 0 or not valid_israeli_id(tz_raw):
            continue
        tz = str(tz_raw).zfill(9)
        out[tz] = _read_block_components(block)
    return out


def _read_block_components(block):
    """
    Parse one employee's block into {component_number: (qty, price, bn)}.
    Walks the packed slot array from SLOT_TABLE_OFFSET, stopping at the
    first all-zero slot. qty/price are floats in whole-shekel units; bn is
    the component's ברוטו/נטו flag as a single char ("ב"/"נ", or "" if none).
    """
    comps = {}
    for i in range(_SLOT_WALK_GUARD):
        off = SLOT_TABLE_OFFSET + i * SLOT_SIZE
        slot = block[off : off + SLOT_SIZE]
        if len(slot) < SLOT_SIZE:
            break
        comp = slot[SLOT_COMP_NUM_OFF]
        qty_raw = struct.unpack_from("<i", slot, SLOT_QTY_OFF)[0]
        price_raw = struct.unpack_from("<i", slot, SLOT_PRICE_OFF)[0]
        if comp == 0 and qty_raw == 0 and price_raw == 0:
            break  # terminator
        if comp == 0:
            continue  # value with no component number -- unmappable, skip
        bn = _decode_bn(slot[SLOT_BN_OFF])
        comps[comp] = (qty_raw / 100.0, price_raw / 100.0, bn)
    return comps


# Keyword families for components that should NOT be carried forward by
# default: vacation (חופש), sick (מחלה), holiday (חג), and maternity (לידה).
# Recreation pay (הבראה) is deliberately NOT here -- it is a real recurring
# payment, not leave. This is only a SUGGESTED default; the operator can add
# or remove components in the UI. Matching is by name (component numbers vary
# per company, so a fixed number list could not generalize).
_DEFAULT_NO_CARRY_KEYWORDS = ("חופש", "מחלה", "חג", "לידה")


def default_no_carry_components(components):
    """
    Given the components list from extract_components() -- each a
    (rechiv_extracted, name, kod_mahk) tuple -- return the set of REAL מיכפל
    component numbers whose names look like leave (vacation/sick/holiday/
    maternity). These are the suggested default "do not carry forward" set.

    The returned numbers are the REAL component numbers as read_prior_month()
    reports them from slot byte [7] -- i.e. `actual` (= rechiv_extracted - 1),
    with the special-cased משכורת block being component 1. These are the same
    keys build_excel()'s `no_carry` is compared against, so leave components
    are correctly excluded.
    """
    out = set()
    for rechiv_extracted, name, _kod in components:
        if any(kw in name for kw in _DEFAULT_NO_CARRY_KEYWORDS):
            # rechiv_extracted == 2 is the משכורת fixed block -> real component 1
            real = 1 if rechiv_extracted == 2 else rechiv_extracted - 1
            out.add(real)
    return out


def extract_employees(path, include_inactive=False):
    """
    Read ACTIVE employees from Q8OVDM26.[company].

    Thin wrapper over _read_employee_records(): keeps employees whose
    קוד הפסקה (the byte at KOD_HAFSAKA_OFFSET within each EMPLOYEE_BLOCK)
    == 0. Inactive/stopped employees (status == 1) are included only when
    include_inactive=True. Unused slots (status == 2) and nameless blocks
    are always excluded.

    Returns (results, invalid_count).
      results       -- list of (tz_str, full_name) sorted by תעודת זהות
      invalid_count -- count of KEPT employees whose תז failed validation
                       (diagnostic only; their tz field is left "")
    """
    results = []
    invalid_ids = 0
    for tz, name, status in _read_employee_records(path):
        if status == STATUS_ACTIVE or (include_inactive and status == STATUS_INACTIVE):
            if not tz:
                invalid_ids += 1
            results.append((tz, name))
    results.sort(key=lambda x: x[0])
    return results, invalid_ids


def extract_employees_with_status(path):
    """
    Read employees split by status -- for UIs that let the operator choose
    which inactive employees (if any) to include in the import template.

    Returns (active, inactive, invalid_count):
      active        -- list of (tz_str, full_name), status 0, sorted by תז
      inactive      -- list of (tz_str, full_name), status 1, sorted by תז
      invalid_count -- employees across BOTH lists whose תז failed
                       validation (their tz field is left "")
    Unused slots (status 2) and nameless blocks are excluded from both.
    """
    active, inactive = [], []
    invalid_ids = 0
    for tz, name, status in _read_employee_records(path):
        if not tz:
            invalid_ids += 1
        if status == STATUS_ACTIVE:
            active.append((tz, name))
        else:  # only ACTIVE / INACTIVE records reach here
            inactive.append((tz, name))
    active.sort(key=lambda x: x[0])
    inactive.sort(key=lambda x: x[0])
    return active, inactive, invalid_ids


def _read_employee_records(path):
    """
    Low-level read of Q8OVDM26.[company]. Returns a list of
    (tz_str, full_name, status) for every block holding a REAL employee --
    a recognized status (active or inactive) AND a non-garbage name. Unused
    pre-allocated slots (status == 2) and nameless/garbage blocks are
    skipped. tz_str is the zero-padded valid ID, or "" if the תז failed
    validation. Order follows the file (callers sort as needed).
    """
    data = open(path, "rb").read()
    records = []
    total_blocks = len(data) // EMPLOYEE_BLOCK
    for b in range(total_blocks):
        block = data[b * EMPLOYEE_BLOCK : (b + 1) * EMPLOYEE_BLOCK]
        status = block[KOD_HAFSAKA_OFFSET]

        # Keep only real employees (active or inactive). Unused slots
        # (status 2) and any other value are skipped.
        if status not in (STATUS_ACTIVE, STATUS_INACTIVE):
            continue

        end = block.index(b"\x00", 9)
        last_name = sanitize_text(
            block[9:end].decode("iso-8859-8", errors="replace").strip()
        )
        end2 = block.index(b"\x00", 29)
        first_name = sanitize_text(
            block[29:end2].decode("iso-8859-8", errors="replace").strip()
        )
        full_name = f"{first_name} {last_name}".strip()

        # A block can carry status 0 yet hold no real name (e.g. trailing
        # padding). Drop nameless blocks, and names that decoded to garbage
        # (U+FFFD survives sanitize_text).
        if not full_name or "\ufffd" in full_name:
            continue

        tz_raw = struct.unpack_from("<I", block, 59)[0]
        tz = str(tz_raw).zfill(9) if valid_israeli_id(tz_raw) else ""
        records.append((tz, full_name, status))

    return records


# ---------------------------------------------------------------------------
# Template builder (@@QD@@ block for Q8SRGL26.000)
# ---------------------------------------------------------------------------


def build_template(template_name, components, deductions=None):
    """
    Build the @@QD@@ text block for insertion into Q8SRGL26.000.

    Returns (template_text, col_map, stats_cols, ded_map).
      col_map    -- list of (actual_code, name, col_כמות, col_מחיר, col_בנ)
      stats_cols -- list of (field_code, label, col_index)
      ded_map    -- list of (deduction_number, name, col_amount, col_month)
                    for each deduction (רכיבי ניכוי רשות); empty if none.

    Deductions are a separate component family (see extract_deductions). In
    the @D block each deduction is emitted as:
        @D<ded_number+1> -2 -2 1     (section separator; field code -2)
        @D<ded_number+1> 25 <col> 1  (amount  -- field code 25, סכום)
        @D<ded_number+1> 26 <col+1> 1 (month  -- field code 26)
    matching the encoding מיכפל itself produced for company 004's מפרעה
    example (@D1 25 .. / @D1 26 ..; the +1 because the @D line uses
    deduction_number + 1, מפרעה=0 -> @D1). Deduction columns are placed
    AFTER the last salary component and BEFORE the stats columns.
    """
    deductions = deductions or []
    lines = []
    lines.append(f"@N{template_name}")
    lines.append("@V90072")
    lines.append("@YB2 C2 A2")
    lines.append("@F4")
    lines.append("@H")
    lines.append("@D-1 1 1 1")
    lines.append("@D1 -1 -1 1")
    lines.append("@D1 7 4 1")  # סכום -> col D=4
    lines.append("@D1 8 3 1")  # ברוטו/נטו -> col C=3
    lines.append("@Mבב")
    lines.append("@M* אין מיפוי *ג")
    lines.append("@Mננ")
    lines.append("@M* אין מיפוי *ע")
    lines.append("@M* אין מיפוי *פ")
    lines.append("@Mקק")
    # רכיב 001 (משכורת) fixed dual-entry at E=5, F=6
    lines.append("@D1 -1 -1 1")
    lines.append("@D1 5 5 1")
    lines.append("@D1 6 6 1")
    lines.append("@D1 8 3 1")

    col_map = []
    col = 7  # remaining components start at G=7
    for rechiv_extracted, name, kod_mahk in components:
        if rechiv_extracted == 2:  # משכורת already covered in fixed block
            continue
        actual = rechiv_extracted - 1  # THE -1 OFFSET -- universally verified
        # Each non-משכורת component now occupies THREE columns: כמות, מחיר, and
        # its own ברוטו/נטו. The flag is per-component in the file (Section 7),
        # so field 8 points at this component's own column -- not the shared
        # column C. (משכורת keeps C as its ב/נ via the fixed block above.)
        col_k, col_m, col_bn = col, col + 1, col + 2
        col_map.append((actual, name, col_k, col_m, col_bn))
        lines.append(f"@D{actual} -1 -1 1")
        lines.append(f"@D{actual} 5 {col_k} 1")
        lines.append(f"@D{actual} 6 {col_m} 1")
        lines.append(f"@D{actual} 8 {col_bn} 1")  # ברוטו/נטו -> own column
        col += 3

    # --- deductions (רכיבי ניכוי רשות): amount + month, after salary,
    #     before the stats block. Field codes 25=amount, 26=month; the @D
    #     line uses deduction_number + 1 (מפרעה=0 -> @D1). ------------------
    ded_map = []
    for ded_number, name, _kod in deductions:
        col_amount, col_month = col, col + 1
        ded_map.append((ded_number, name, col_amount, col_month))
        d = ded_number + 1
        lines.append(f"@D{d} -2 -2 1")
        lines.append(f"@D{d} 25 {col_amount} 1")
        lines.append(f"@D{d} 26 {col_month} 1")
        col += 2

    stats = [
        (10, "ימי עבודה משולמים"),
        (14, "ימי עבודה בפועל"),
        (15, "שעות בפועל"),
        (16, "תקן ימים"),
        (17, "תקן שעות"),
        (18, "חופשה"),
        (19, "מחלה"),
    ]
    stats_cols = []
    for field, label in stats:
        lines.append(f"@D-1 {field} {col} 1")
        stats_cols.append((field, label, col))
        col += 1

    return "\n".join(lines) + "\n", col_map, stats_cols, ded_map


# ---------------------------------------------------------------------------
# Excel builder
# ---------------------------------------------------------------------------


def build_excel(
    company,
    components,
    col_map,
    stats_cols,
    employees,
    out_path,
    year,
    month,
    prior_month: "dict | None" = None,
    no_carry: "set | None" = None,
    ded_map=None,
    prior_deductions: "dict | None" = None,
):
    """
    Generate the .xlsx import template.

    Rows 1-2 : metadata (חברה / שנת מס / חודש דיווח)
    Row 4    : column headers
    Row 5+   : employees (תעודת זהות pre-filled in col A, name in col B)

    Carry-forward (optional):
      prior_month -- dict from read_prior_month(): {tz: {component_number:
                     (qty, price, bn)}}. When given, each employee's row is
                     pre-filled with last month's qty/price for every
                     component that maps to a column, plus that component's
                     ברוטו/נטו flag (ב/נ) in its own column. Components are
                     joined by REAL מיכפל number (= col_map `actual`, and the
                     fixed משכורת block = component 1, whose ב/נ is column C).
      no_carry    -- iterable of REAL מיכפל component numbers NOT to carry
                     forward (e.g. vacation/sick). Defaults to none excluded.

    Deductions (רכיבי ניכוי רשות, optional):
      ded_map          -- list of (deduction_number, name, col_amount,
                          col_month) from build_template(). Adds an amount
                          column and a month column per deduction (headers
                          = deduction name + " סכום" / " חודש").
      prior_deductions -- dict from read_prior_month_deductions():
                          {tz: {deduction_number: amount}}. When given,
                          each employee's deduction AMOUNT cell is pre-filled
                          (carry-forward). Month cells are left blank for the
                          operator to fill. no_carry is NOT applied to
                          deductions (leave-type exclusions are a salary
                          concept).

    New-hire highlighting: when prior_month is given, any employee whose תז
    is NOT a key in prior_month had no prior-month record (a new hire). Their
    תעודת זהות (col A) and שם (col B) cells are filled YELLOW so the operator
    can see at a glance which rows start blank because there was nothing to
    carry. Employees present last month but with no activity ({} ) are NOT
    highlighted -- they existed, they just had no lines.

    Returns a dict:
        {"out_path": str, "new_hires": [(tz, name), ...],
         "no_activity": [(tz, name), ...], "carried": int}
    where `carried` is the total number of (employee, component) cells filled
    (salary + deductions).
    """
    no_carry = set(no_carry or ())
    ded_map = ded_map or []
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None  # a fresh Workbook always has an active sheet;
    #                        narrows openpyxl's Worksheet|None for type checkers
    ws.title = "ייבוא משכורות"
    ws.sheet_view.rightToLeft = True

    # --- styles ---------------------------------------------------------
    bold = Font(bold=True, name="Arial", size=10)
    normal = Font(name="Arial", size=10)
    header_fill = PatternFill("solid", start_color="D9E1F2")
    comp_fill = PatternFill("solid", start_color="E2EFDA")
    stats_fill = PatternFill("solid", start_color="FFF2CC")
    ded_fill = PatternFill("solid", start_color="FBE2D5")  # deduction columns
    rechiv1_fill = PatternFill("solid", start_color="FCE4D6")
    emp_fill = PatternFill("solid", start_color="F2F2F2")
    newhire_fill = PatternFill("solid", start_color="FFFF00")  # new-hire תז/שם marker
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="right", vertical="center")
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style(cell, font=normal, fill=None, align=center):
        cell.font = font
        if fill:
            cell.fill = fill
        cell.alignment = align
        cell.border = border

    def safe_set(cell, value):
        """Set a cell value, sanitizing strings as a final safety net."""
        if isinstance(value, str):
            value = sanitize_text(value)
        cell.value = value
        return cell

    # --- rows 1-2: metadata ---------------------------------------------
    style(safe_set(ws["A1"], "חברה"), font=bold, fill=header_fill)
    style(safe_set(ws["B1"], "שנת מס"), font=bold, fill=header_fill)
    style(safe_set(ws["C1"], "חודש דיווח"), font=bold, fill=header_fill)

    style(safe_set(ws["A2"], int(company)))
    style(safe_set(ws["B2"], int(year)))  # read by @YB2 C2 A2
    style(safe_set(ws["C2"], int(month)))  # read by @YB2 C2 A2

    # --- row 4: column headers ------------------------------------------
    for col_idx, label in [
        (1, "מספר עובד"),
        (2, "שם עובד"),
        (3, "ברוטו/נטו"),
        (4, "סכום"),
    ]:
        style(safe_set(ws.cell(4, col_idx), label), font=bold, fill=header_fill)

    first_comp_name = components[0][1] if components else "משכורת"
    style(
        safe_set(ws.cell(4, 5), f"{first_comp_name}\nכמות"),
        font=bold,
        fill=rechiv1_fill,
    )
    style(
        safe_set(ws.cell(4, 6), f"{first_comp_name}\nמחיר"),
        font=bold,
        fill=rechiv1_fill,
    )

    for actual, name, col_k, col_m, col_bn in col_map:
        style(safe_set(ws.cell(4, col_k), f"{name}\nכמות"), font=bold, fill=comp_fill)
        style(safe_set(ws.cell(4, col_m), f"{name}\nמחיר"), font=bold, fill=comp_fill)
        style(
            safe_set(ws.cell(4, col_bn), f"{name}\nברוטו/נטו"),
            font=bold,
            fill=comp_fill,
        )

    # deduction headers: amount + month per deduction (after salary, before stats)
    for ded_number, name, col_amount, col_month in ded_map:
        style(
            safe_set(ws.cell(4, col_amount), f"{name}\nסכום"), font=bold, fill=ded_fill
        )
        style(
            safe_set(ws.cell(4, col_month), f"{name}\nחודש"), font=bold, fill=ded_fill
        )

    for field, label, col_idx in stats_cols:
        style(safe_set(ws.cell(4, col_idx), label), font=bold, fill=stats_fill)

    # --- component-number -> (qty_col, price_col) lookup for carry-forward --
    # Join key is the REAL מיכפל number that read_prior_month() reads from slot
    # byte [7]. That value equals col_map's `actual` DIRECTLY (verified against
    # company 004's real exported template: רכב actual=2 -> slot 2, טלפון
    # actual=3 -> slot 3, ש.נוס125% actual=7 -> slot 7, בונוס actual=18 -> slot
    # 18). The fixed משכורת block is real component 1 at E/F (cols 5/6).
    #
    # NOTE: an earlier version used `actual + 1` here, derived from a synthetic
    # test that mismodeled the משכורת/רכב boundary; it shifted every carried
    # value one component to the left. Do not reintroduce the +1.
    comp_to_cols = {1: (5, 6, 3)}  # משכורת: כמות E=5, מחיר F=6, ברוטו/נטו C=3
    for actual, name, col_k, col_m, col_bn in col_map:
        comp_to_cols[actual] = (col_k, col_m, col_bn)

    # deduction-number -> amount column (month column left for the operator)
    ded_to_col = {dn: col_amount for dn, _name, col_amount, _col_month in ded_map}

    # --- rows 5+: employees ---------------------------------------------
    new_hires = []
    no_activity = []
    carried = 0
    for row_idx, (tz, full_name) in enumerate(employees, 5):
        # "new hire" is judged on salary prior_month (the primary signal);
        # deductions are carried independently below regardless.
        is_new_hire = prior_month is not None and tz not in prior_month
        id_fill = newhire_fill if is_new_hire else emp_fill

        style(safe_set(ws.cell(row_idx, 1), tz), fill=id_fill)
        style(safe_set(ws.cell(row_idx, 2), full_name), fill=id_fill, align=left)

        # salary carry-forward
        if prior_month is not None and not is_new_hire:
            emp_prior = prior_month[tz]
            if not emp_prior:
                no_activity.append((tz, full_name))
            else:
                for comp_num, (qty, price, bn) in emp_prior.items():
                    if comp_num in no_carry:
                        continue
                    cols = comp_to_cols.get(comp_num)
                    if cols is None:
                        continue  # component has no column in this template
                    col_k, col_m, col_bn = cols
                    style(safe_set(ws.cell(row_idx, col_k), qty), fill=emp_fill)
                    style(safe_set(ws.cell(row_idx, col_m), price), fill=emp_fill)
                    if bn:
                        style(safe_set(ws.cell(row_idx, col_bn), bn), fill=emp_fill)
                    carried += 1
        elif is_new_hire:
            new_hires.append((tz, full_name))

        # deduction carry-forward (independent of salary lines; amount only,
        # month left blank for the operator)
        if prior_deductions is not None and ded_to_col:
            emp_ded = prior_deductions.get(tz, {})
            for ded_num, amount in emp_ded.items():
                col_amount = ded_to_col.get(ded_num)
                if col_amount is None:
                    continue  # deduction has no column in this template
                style(safe_set(ws.cell(row_idx, col_amount), amount), fill=emp_fill)
                carried += 1

    # --- column widths --------------------------------------------------
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 10
    for actual, name, col_k, col_m, col_bn in col_map:
        ws.column_dimensions[col_letter(col_k)].width = 10
        ws.column_dimensions[col_letter(col_m)].width = 10
        ws.column_dimensions[col_letter(col_bn)].width = 9
    for ded_number, name, col_amount, col_month in ded_map:
        ws.column_dimensions[col_letter(col_amount)].width = 11
        ws.column_dimensions[col_letter(col_month)].width = 8
    for field, label, col_idx in stats_cols:
        ws.column_dimensions[col_letter(col_idx)].width = 13

    ws.row_dimensions[4].height = 45
    ws.freeze_panes = "E5"
    wb.save(out_path)
    return {
        "out_path": out_path,
        "new_hires": new_hires,
        "no_activity": no_activity,
        "carried": carried,
    }


# ---------------------------------------------------------------------------
# Q8SRGL26.000 backup + insertion
# ---------------------------------------------------------------------------


def regenerate_and_insert_template(srgl_path, template_name, template_text, company):
    """
    Idempotently (re)place a company's template block in Q8SRGL26.000:
    remove any existing block with this exact name first, then insert
    the new template_text. Always backs up before any modification.

    Returns a dict:
        {
            "removed":       bool,           # was an old block found+removed?
            "remove_backup": str or None,     # backup path from the removal step
            "insert_backup": str,             # backup path from the insertion step
            "removed_block": str or None,     # the old block's text, for review
        }

    Raises ValueError if the template name is found more than once
    BEFORE touching the file at all (refuses to guess).
    """
    existing = list_existing_templates(srgl_path)
    matches = [n for n in existing if n == template_name]

    if len(matches) > 1:
        raise ValueError(
            f"Template {template_name!r} found {len(matches)} times in "
            f"{srgl_path} -- refusing to regenerate until the file is "
            f"de-duplicated by hand."
        )

    result = {
        "removed": False,
        "remove_backup": None,
        "insert_backup": None,
        "removed_block": None,
    }

    if len(matches) == 1:
        remove_backup, removed_block = remove_template(srgl_path, template_name)
        result["removed"] = True
        result["remove_backup"] = remove_backup
        result["removed_block"] = removed_block

    insert_backup = backup_and_insert(srgl_path, template_text, company)
    result["insert_backup"] = insert_backup

    return result


def remove_template(srgl_path, template_name):
    """
    Back up Q8SRGL26.000, then remove the single @N<template_name> block
    (from its @N line up to the next @N line or @@XL@@, whichever comes
    first) from the file.

    Raises ValueError BEFORE touching the original file if template_name
    is not found, or if it is found more than once (ambiguous -- this
    function will not guess which one to remove).

    Returns (backup_path, removed_block_text) on success.
    """
    if not os.path.exists(srgl_path):
        raise FileNotFoundError(f"{srgl_path} not found")

    with open(srgl_path, "rb") as f:
        content = f.read()

    text = content.decode("iso-8859-8", errors="replace")
    lines = text.splitlines(keepends=True)

    target_line = f"@N{template_name}"
    match_indices = [
        i for i, line in enumerate(lines) if line.rstrip("\r\n") == target_line
    ]

    if len(match_indices) == 0:
        raise ValueError(f"Template {template_name!r} not found -- nothing removed.")
    if len(match_indices) > 1:
        raise ValueError(
            f"Template {template_name!r} found {len(match_indices)} times -- "
            f"refusing to guess which to remove."
        )

    start = match_indices[0]
    end = len(lines)
    for j in range(start + 1, len(lines)):
        s = lines[j].rstrip("\r\n")
        if s.startswith("@N") or s.startswith("@@XL@@"):
            end = j
            break

    removed_block = "".join(lines[start:end])

    # --- backup BEFORE any modification, same pattern as backup_and_insert ---
    company_hint = "".join(ch for ch in template_name if ch.isdigit()) or "removal"
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = f"{srgl_path}.bak-{company_hint}-remove-{timestamp}"

    with open(backup_path, "wb") as f:
        f.write(content)
    with open(backup_path, "rb") as f:
        if f.read() != content:
            raise IOError("Backup verification failed -- original file left untouched.")
    # ---------------------------------------------------------------------------

    new_text = "".join(lines[:start]) + "".join(lines[end:])
    new_content = new_text.encode("iso-8859-8")

    with open(srgl_path, "wb") as f:
        f.write(new_content)

    return backup_path, removed_block


def backup_and_insert(srgl_path, template_text, company):
    """
    Back up Q8SRGL26.000, then insert template_text (the @@QD@@ block)
    immediately before the @@XL@@ marker.

    Backup naming:
        {srgl_path}.bak-{company zero-padded to 3 digits}-{YYYYMMDD_HHMMSS}

    All validation happens BEFORE the original file is touched. The backup
    is written and verified first; the marker is checked next; only then
    is the original file overwritten. Returns the backup path on success.
    """
    if not os.path.exists(srgl_path):
        raise FileNotFoundError(f"{srgl_path} not found")

    company_padded = str(company).zfill(3)
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = f"{srgl_path}.bak-{company_padded}-{timestamp}"

    with open(srgl_path, "rb") as f:
        content = f.read()

    # Backup BEFORE any modification attempt.
    with open(backup_path, "wb") as f:
        f.write(content)

    # Verify the backup actually matches before proceeding.
    with open(backup_path, "rb") as f:
        if f.read() != content:
            raise IOError("Backup verification failed -- original file left untouched.")

    marker = b"@@XL@@"
    pos = content.find(marker)
    if pos == -1:
        raise ValueError("Marker @@XL@@ not found -- refusing to modify file.")

    template_encoded = template_text.encode("iso-8859-8")
    new_content = content[:pos] + template_encoded + b"\n" + content[pos:]

    with open(srgl_path, "wb") as f:
        f.write(new_content)

    return backup_path
